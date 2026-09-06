"""Bounded position import with preview, explicit replacement, and conflict-safe undo."""
from __future__ import annotations

import csv
import hashlib
import io
import re
import zipfile
from datetime import UTC, datetime, timedelta
from decimal import Decimal, InvalidOperation
from uuid import uuid4

from pydantic import Field, field_validator
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import Holding, Instrument
from app.services.holding_service import HoldingService
from app.workspace.config import workspace_settings
from app.workspace.jobs import WorkspaceError, lock_owner, owner_scope, utc
from app.workspace.models import WorkspaceImportBatch
from app.workspace.protocol import Hash, StrictModel, content_hash, safe_text

ALIASES = {
    "ts_code": {"ts_code", "code", "symbol", "代码", "基金代码", "证券代码"},
    "shares": {"shares", "quantity", "份额", "持仓份额", "持有份额", "持有数量", "证券数量"},
    "cost_price": {"cost_price", "unit_cost", "成本价", "单位成本", "每份成本", "持仓成本价"},
}


class CandidateEdit(StrictModel):
    row_index: int = Field(ge=1, le=500)
    ts_code: str = Field(max_length=32)
    shares: str = Field(max_length=32)
    cost_price: str = Field(max_length=32)
    selected: bool = True

    @field_validator("ts_code", "shares", "cost_price")
    @classmethod
    def safe(cls, value: str) -> str:
        return safe_text(value)


class ImportEdit(StrictModel):
    expected_hash: Hash
    candidates: list[CandidateEdit] = Field(min_length=1, max_length=500)


class ImportConfirm(StrictModel):
    expected_hash: Hash


class ManualPreview(StrictModel):
    candidates: list[CandidateEdit] = Field(min_length=1, max_length=200)


def decimal_value(value: str) -> str:
    value = str(value).strip()
    if not re.fullmatch(r"\d{1,12}(?:\.\d{1,8})?", value):
        raise ValueError("必须填写非负十进制数，不能使用公式、百分比或科学计数法")
    parsed = Decimal(value)
    if not parsed.is_finite() or parsed > Decimal("1000000000000"):
        raise ValueError("数值超出范围")
    return str(parsed)


def _headers(values) -> dict[str, int]:
    headers = [str(value or "").strip().lower().replace("\ufeff", "") for value in values]
    mapping: dict[str, int] = {}
    for name, aliases in ALIASES.items():
        indexes = [i for i, value in enumerate(headers) if value in aliases]
        if len(indexes) != 1:
            raise WorkspaceError(422, f"import_header_{name}_required_unambiguous")
        mapping[name] = indexes[0]
    return mapping


def parse_file(data: bytes, extension: str) -> list[dict]:
    cfg = workspace_settings()
    if len(data) > cfg.import_max_bytes:
        raise WorkspaceError(413, "import_file_too_large")
    if extension == ".csv":
        try:
            text = data.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = data.decode("gb18030")
            except UnicodeDecodeError:
                raise WorkspaceError(422, "import_encoding_invalid") from None
        try:
            reader = csv.reader(io.StringIO(text), strict=True)
            header = next(reader)
            mapping = _headers(header)
            raw = []
            for values in reader:
                if not any(str(value).strip() for value in values):
                    continue
                if len(raw) >= cfg.import_max_rows or len(values) > 200:
                    raise WorkspaceError(422, "import_row_or_column_limit")
                raw.append({name: str(values[index]).strip() if index < len(values) else "" for name, index in mapping.items()})
        except (csv.Error, StopIteration):
            raise WorkspaceError(422, "import_csv_invalid") from None
    elif extension == ".xlsx":
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                infos = archive.infolist()
                if len(infos) > 200 or sum(info.file_size for info in infos) > 10_000_000 or any(info.file_size > 2_000_000 or info.file_size > max(info.compress_size, 1) * 1000 or ".." in info.filename.split("/") or "vbaProject" in info.filename or "externalLink" in info.filename for info in infos):
                    raise WorkspaceError(422, "import_xlsx_unsafe_archive")
            import openpyxl
            book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=False, keep_links=False)
            try:
                if len(book.worksheets) != 1:
                    raise WorkspaceError(422, "import_requires_single_sheet")
                sheet = book.worksheets[0]
                if sheet.max_column > 200 or sheet.max_row > cfg.import_max_rows + 1:
                    raise WorkspaceError(422, "import_row_or_column_limit")
                rows = sheet.iter_rows()
                header = next(rows)
                mapping = _headers([cell.value for cell in header])
                raw = []
                for cells in rows:
                    if not any(cell.value is not None for cell in cells):
                        continue
                    if len(raw) >= cfg.import_max_rows:
                        raise WorkspaceError(422, "import_row_limit")
                    values = {}
                    for name, index in mapping.items():
                        cell = cells[index]
                        if cell.data_type == "f":
                            raise WorkspaceError(422, "import_formulas_forbidden")
                        value = cell.value
                        if isinstance(value, float) and value.is_integer():
                            value = int(value)
                        values[name] = str(value) if value is not None else ""
                    raw.append(values)
            finally:
                book.close()
        except WorkspaceError:
            raise
        except ImportError:
            raise WorkspaceError(503, "xlsx_dependency_unavailable") from None
        except Exception:
            raise WorkspaceError(422, "import_xlsx_invalid") from None
    else:
        raise WorkspaceError(415, "import_csv_or_xlsx_required")
    if not raw:
        raise WorkspaceError(422, "import_empty")
    return [{**row, "row_index": index + 1, "selected": True} for index, row in enumerate(raw)]


def validate_candidates(db: Session, candidates: list[dict]) -> list[dict]:
    all_codes = {str(row.get("ts_code", "")).strip().upper() for row in candidates}
    symbols = {code for code in all_codes if re.fullmatch(r"\d{6}", code)}
    instruments = db.scalars(select(Instrument).where(Instrument.kind.in_(("ETF", "LOF")), (Instrument.ts_code.in_(all_codes) | Instrument.symbol.in_(symbols)))).all()
    by_code = {row.ts_code: row for row in instruments}
    by_symbol: dict[str, list] = {}
    for row in instruments:
        by_symbol.setdefault(row.symbol, []).append(row)
    result, seen = [], set()
    for candidate in candidates:
        code = str(candidate.get("ts_code", "")).strip().upper()
        errors = []
        inst = by_code.get(code)
        if inst is None and len(by_symbol.get(code, [])) == 1:
            inst = by_symbol[code][0]
        if inst is None:
            errors.append("代码未能唯一匹配已同步场内 ETF/LOF，请保留六位前导零和交易所")
        matched = inst.ts_code if inst else code
        values = {}
        for key in ("shares", "cost_price"):
            try:
                values[key] = decimal_value(str(candidate.get(key, "")))
            except ValueError as exc:
                values[key] = str(candidate.get(key, ""))[:32]
                errors.append(f"{key}: {exc}")
        if candidate.get("selected", True) and matched in seen:
            errors.append("重复代码：请合并份额与每份成本，或取消选择重复行")
        if candidate.get("selected", True):
            seen.add(matched)
        result.append({"row_index": int(candidate["row_index"]), "ts_code": matched, "name": inst.name if inst else None, **values, "selected": bool(candidate.get("selected", True)), "errors": errors, "action": "replace_after_confirmation"})
    return result


def preview(db: Session, data: bytes, extension: str, user_id: int | None) -> WorkspaceImportBatch:
    return preview_rows(db, parse_file(data, extension), hashlib.sha256(data).hexdigest(), extension.removeprefix("."), user_id)


def preview_rows(db: Session, candidates: list[dict], source_hash: str, source_kind: str, user_id: int | None) -> WorkspaceImportBatch:
    scope = owner_scope(user_id)
    lock_owner(db, scope)
    existing = db.scalar(select(WorkspaceImportBatch).where(WorkspaceImportBatch.owner_scope == scope, WorkspaceImportBatch.source_hash == source_hash))
    if existing:
        return existing
    row = WorkspaceImportBatch(batch_id=uuid4().hex, user_id=user_id, owner_scope=scope, source_hash=source_hash, source_kind=source_kind, candidates_json=validate_candidates(db, candidates), expires_at=datetime.now(UTC) + timedelta(hours=2))
    db.add(row)
    db.flush()
    return row


def owned_batch(db: Session, batch_id: str, scope: str) -> WorkspaceImportBatch:
    row = db.scalar(select(WorkspaceImportBatch).where(WorkspaceImportBatch.batch_id == batch_id, WorkspaceImportBatch.owner_scope == scope).with_for_update())
    if row is None:
        raise WorkspaceError(404, "import_batch_not_found")
    return row


def view(row: WorkspaceImportBatch) -> dict:
    return {"batch_id": row.batch_id, "status": row.status, "source_kind": row.source_kind, "expires_at": row.expires_at.isoformat(), "candidates": row.candidates_json, "candidate_hash": content_hash(row.candidates_json), "confirmed_at": row.confirmed_at.isoformat() if row.confirmed_at else None, "mode": "replace_selected_instruments", "note": "成本为每份成本，份额为当前总份额；确认会替换选中基金的现有记录，不进行加法。"}


def edit(db: Session, row: WorkspaceImportBatch, payload: ImportEdit) -> WorkspaceImportBatch:
    if row.status != "preview" or utc(row.expires_at) < datetime.now(UTC):
        raise WorkspaceError(409, "import_preview_inactive")
    if content_hash(row.candidates_json) != payload.expected_hash:
        raise WorkspaceError(409, "import_preview_changed")
    indexes = [item.row_index for item in payload.candidates]
    if len(set(indexes)) != len(indexes) or set(indexes) != {item["row_index"] for item in row.candidates_json}:
        raise WorkspaceError(422, "import_candidate_identity_mismatch")
    row.candidates_json = validate_candidates(db, [item.model_dump() for item in payload.candidates])
    db.flush()
    return row


def holding_state(db: Session, user_id: int | None, codes: list[str]) -> dict:
    pairs = db.execute(select(Holding, Instrument.ts_code).join(Instrument, Holding.instrument_id == Instrument.id).where(Holding.user_id == user_id, Instrument.ts_code.in_(codes)).with_for_update()).all()
    result = dict.fromkeys(codes)
    for holding, code in pairs:
        result[code] = {"shares": str(holding.shares.normalize()), "cost_price": str(holding.cost_price.normalize()), "target_weight": holding.target_weight, "notes": holding.notes}
    return result


def confirm(db: Session, row: WorkspaceImportBatch, expected_hash: str) -> WorkspaceImportBatch:
    if content_hash(row.candidates_json) != expected_hash:
        raise WorkspaceError(409, "import_preview_changed")
    if row.status == "confirmed":
        return row
    if row.status != "preview" or utc(row.expires_at) < datetime.now(UTC):
        raise WorkspaceError(409, "import_preview_inactive")
    current = validate_candidates(db, row.candidates_json)
    selected = [item for item in current if item["selected"]]
    if not selected or any(item["errors"] for item in selected):
        raise WorkspaceError(422, "import_candidates_require_correction")
    lock_owner(db, row.owner_scope)
    codes = [item["ts_code"] for item in selected]
    before = holding_state(db, row.user_id, codes)
    service = HoldingService()
    for item in selected:
        old = before[item["ts_code"]] or {}
        service.upsert(db, user_id=row.user_id, ts_code=item["ts_code"], shares=Decimal(item["shares"]), cost_price=Decimal(item["cost_price"]), target_weight=old.get("target_weight"), notes=old.get("notes"))
    row.before_json, row.after_hash = before, content_hash(holding_state(db, row.user_id, codes))
    row.status, row.confirmed_at = "confirmed", datetime.now(UTC)
    db.flush()
    return row


def undo(db: Session, row: WorkspaceImportBatch) -> WorkspaceImportBatch:
    if row.status == "undone":
        return row
    if row.status != "confirmed" or row.before_json is None:
        raise WorkspaceError(409, "import_not_confirmed")
    lock_owner(db, row.owner_scope)
    if content_hash(holding_state(db, row.user_id, list(row.before_json))) != row.after_hash:
        raise WorkspaceError(409, "holdings_changed_after_import_cannot_undo")
    service = HoldingService()
    for code, old in row.before_json.items():
        if old is None:
            service.delete(db, code, user_id=row.user_id)
        else:
            service.upsert(db, user_id=row.user_id, ts_code=code, shares=Decimal(old["shares"]), cost_price=Decimal(old["cost_price"]), target_weight=old.get("target_weight"), notes=old.get("notes"))
    row.status = "undone"
    db.flush()
    return row
